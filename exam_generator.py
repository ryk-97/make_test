import openpyxl as xl
from datetime import datetime
import random
import math
import sympy as sp

# 現在時刻を取得
now = datetime.now()
formatted_now = now.strftime('%Y%m%d-%H%M%S')

# ファイルを読み込み、新しいシートを作成
wb = xl.load_workbook('exam.xlsx')
ws = wb.create_sheet(title='新しいシート' + str(formatted_now))

# シートに問題を書き込む
for i in range(1, 31):
    num1 = random.randint(-10, 10)
    num2 = random.randint(-10, 10)

    mode = random.choice(['+', '-', '×', '÷'])
    if mode == '+':
        result = num1 + num2
    elif mode == '-':
        result = num1 - num2
    elif mode == '×':
        result = num1 * num2
    elif mode == '÷':
        result = sp.Rational(num1, num2)
    else:
        result = 0

    target_cell = ws.cell(i, 1)
    target_cell.value = f'({num1}) {mode} ({num2})'

    target_cell = ws.cell(i, 2)
    target_cell.value = str(result)

for i in range(1, 31):
    x = sp.symbols('x')

    num1 = random.randint(-10, 10)
    num2 = random.randint(-10, 10)
    num3 = random.randint(-10, 10)

    while num1 == 0 or num2 == 0 or num3 == 0:
        num1 = random.randint(-10, 10)
        num2 = random.randint(-10, 10)
        num3 = random.randint(-10, 10)

    quation = sp.Eq(num1 * x + num2,num3)
    result = sp.solve(quation, x)

    target_cell = ws.cell(i, 3)
    target_cell.value = f'{num1}x + {num2} = {num3}'

    target_cell = ws.cell(i, 4)
    target_cell.value = str(result)[1:-1]

for i in range(1, 31):
    x = sp.symbols('x')

    num1 = random.randint(-10, 10)
    num2 = random.randint(-10, 10)
    num3 = random.randint(-10, 10)

    while num1 == 0 or num2 == 0 or num3 == 0:
        num1 = random.randint(-10, 10)
        num2 = random.randint(-10, 10)
        num3 = random.randint(-10, 10)

    expanded_expression = sp.expand(num1 * (x + num2) + num3)
    result = sp.solve(expanded_expression, x)

    target_cell = ws.cell(i, 5)
    target_cell.value = f'{num1}(x + {num2}) +({num3})'

    target_cell = ws.cell(i, 6)
    target_cell.value = str(expanded_expression)

# シートを保存
wb.save('exam.xlsx')