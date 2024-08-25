import openpyxl as px
from datetime import datetime
import random
import math
import sympy as sp
import re

# mode1は、-の時に括弧をつける　mode2は、+の時に+をつける mode3は、1の時に数字を表示しない
def f_n(num, mode):
    if mode == 3:
        if num == 1 or num == "1":
            return ''
        elif num == -1 or num == "-1":
            return '-'
        return str(num)
    if num < 0 and mode == 1:
        return f'({num})'
    elif num >= 0 and mode == 2:
        return f'+{num}'
    return str(num)
def simplify_quadratic(equation):
    # 正規表現を使って係数を抽出
    match = re.match(r"x\^2\s*([\+\-]?\d*)x\s*([\+\-]?\d*)=0", equation)
    if not match:
        return "Invalid equation format"
    
    # 係数を取得
    a = 1  # x^2 の係数は常に 1
    b = int(match.group(1)) if match.group(1) else 0
    c = int(match.group(2)) if match.group(2) else 0
    
    # 整理された式を作成
    simplified_equation = "x^2"
    if b != 0:
        simplified_equation += f"{'+' if b > 0 else ''}{b}x"
    if c != 0:
        simplified_equation += f"{'+' if c > 0 else ''}{c}"
    
    # 不要なスペースを削除し、+1x を +x に、-1x を -x に置き換え
    simplified_equation = simplified_equation.replace("  ", " ").replace("+1x", "+x").replace("-1x", "-x")
    
    return simplified_equation
"""_summary_
    - 数学の問題を作成する
    - modeによって、問題の種類を指定

    - mode: basic, equation, quadratic, fraction
        - basic: 四則演算
        - equation: 一次方程式
        - quadratic: 二次方程式
        - fraction: 分数

    return:
    - expression: 問題
    - result: 答え
"""
def creat_math_problem(mode):
    num1 = random.randint(1, 10) * random.choice([-1, 1])
    num2 = random.randint(1, 10) * random.choice([-1, 1])
    num3 = random.randint(1, 10) * random.choice([-1, 1])
    num4 = random.randint(1, 10) * random.choice([-1, 1])

    if mode == 'basic':
        basic_type = random.choice(['plus', 'minus', 'multiply', 'divide'])

        if basic_type == 'plus':
            expression = f'{f_n(num1, 1)}+{f_n(num2, 1)}'
            result = num1 + num2
        elif basic_type == 'minus':
            expression = f'{f_n(num1, 1)}-{f_n(num2, 1)}'
            result = num1 - num2
        elif basic_type == 'multiply':
            expression = f'{f_n(num1, 1)}x{f_n(num2, 1)}'
            result = num1 * num2
        elif basic_type == 'divide':
            expression = f'{f_n(num1, 1)}/{f_n(num2, 1)}'
            result = sp.Rational(num1, num2)
    elif mode == 'equation':
        equation_type = random.choice(['type1', 'type2', 'type3'])
        x = sp.symbols('x')

        if equation_type == 'type1':
            # ax + b = cx + d
            expression = f'{f_n(num1,3)}x{f_n(num2,2)}={f_n(num3,3)}x{f_n(num4,2)}'
            result = f'x={str(sp.solve(sp.Eq(num1 * x + num2, num3 * x + num4), x))[1:-1]}'
        elif equation_type == 'type2':
            # ax + b = c
            expression = f'{f_n(num1,3)}x{f_n(num2,2)}={num3}'
            result = f'x={str(sp.solve(sp.Eq(num1 * x + num2, num3), x))[1:-1]}'
        elif equation_type == 'type3':
            # ax = bx + c
            expression = f'{f_n(num1,3)}x={f_n(num2,3)}x{f_n(num3,2)}'
            result = f'x={str(sp.solve(sp.Eq(num1 * x, num2 * x + num3), x))[1:-1]}'
    elif mode == 'quadratic':
        quadratic_type = random.choice(['type1', 'type2', 'type3'])
        if quadratic_type == 'type1':
            b = num1 + num2
            c = num1 * num2

            expression = simplify_quadratic(f'x^2{f_n(b, 2)}x{f_n(c, 2)}=0')
            result = f'x={num1*-1},{num2*-1}'
        elif quadratic_type == 'type2':
            expression = f'x^2{(num1**2)*-1}=0'
            result = f'x={num1},{num1*-1}'
        elif quadratic_type == 'type3':
            x = sp.symbols('x')
            expression = str(sp.expand((x + num1)**2)).replace("**", "^").replace("*", "").replace(" ", "")
            result = f'x={num1*-1}'
    elif mode == 'fraction':
        fraction1 = sp.Rational(1, num1)
        fraction2 = sp.Rational(1, num2)

        result = fraction1 + fraction2
        expression = f'(1/{num1})+(1/{num2})'

    return expression, result

# Get the current time
now = datetime.now()
formatted_now = now.strftime('%Y%m%d%H%M%S')

# Load the file and create a new sheet
wb = px.load_workbook('exam.xlsx')
ws = wb.create_sheet(title='EXAM' + str(formatted_now))

for i in range(1, 101):
    if i <= 25:
        expression, result = creat_math_problem('basic')
        target_cell = ws.cell(i, 1)
        target_cell.value = expression
    elif i <= 50:
        expression, result = creat_math_problem('equation')
        target_cell = ws.cell(i-25, 3)
        target_cell.value = expression
    elif i <= 75:
        expression, result = creat_math_problem('quadratic')
        target_cell = ws.cell(i-50, 5)
        target_cell.value = expression
    else:
        expression, result = creat_math_problem('fraction')
        target_cell = ws.cell(i-75, 7)
        target_cell.value = expression

    if i <= 8:
        target_cell = ws.cell(26, i)
        target_cell.value = str(result)
    elif i <= 16:
        target_cell = ws.cell(27, i-8)
        target_cell.value = str(result)
    elif i <= 24:
        target_cell = ws.cell(28, i-16)
        target_cell.value = str(result)
    elif i <= 32:
        target_cell = ws.cell(29, i-24)
        target_cell.value = str(result)
    elif i <= 40:
        target_cell = ws.cell(30, i-32)
        target_cell.value = str(result)
    elif i <= 48:
        target_cell = ws.cell(31, i-40)
        target_cell.value = str(result)
    elif i <= 56:
        target_cell = ws.cell(32, i-48)
        target_cell.value = str(result)
    elif i <= 64:
        target_cell = ws.cell(33, i-56)
        target_cell.value = str(result)
    elif i <= 72:
        target_cell = ws.cell(34, i-64)
        target_cell.value = str(result)
    elif i <= 80:
        target_cell = ws.cell(35, i-72)
        target_cell.value = str(result)
    elif i <= 88:
        target_cell = ws.cell(36, i-80)
        target_cell.value = str(result)
    elif i <= 96:
        target_cell = ws.cell(37, i-88)
        target_cell.value = str(result)
    elif i <= 104:
        target_cell = ws.cell(38, i-96)
        target_cell.value = str(result)

ws.insert_rows(26)

for i in range(1,101):
    if i <= 25:
        target_cell = ws.cell(i, 2)
        target_cell.value = "_________________"
        ws.column_dimensions["B"].width = 15
    elif i <= 50:
        target_cell = ws.cell(i-25, 4)
        target_cell.value = "_________________"
        ws.column_dimensions["D"].width = 15
    elif i <= 75:
        target_cell = ws.cell(i-50, 6)
        target_cell.value = "_________________"
        ws.column_dimensions["F"].width = 15
    else:
        target_cell = ws.cell(i-75, 8)
        target_cell.value = "_________________"
        ws.column_dimensions["H"].width = 15

ws.insert_rows(1)
ws.insert_rows(1)

ws.merge_cells('A1:H1')
ws['A1'] = '数学基礎計算'
ws.merge_cells('A2:H2')
ws['A2'] = f'名前:雪丸里空 日付:{now.strftime("%Y年%m月%d日")} タイム:______分______秒 点数:______点'

wb.save('exam.xlsx')